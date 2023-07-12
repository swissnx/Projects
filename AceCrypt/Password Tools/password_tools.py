
import string
import secrets
import zxcvbn
import passlib.hash
from datetime import datetime as dt
from cryptography.fernet import Fernet
from openpyxl import Workbook, load_workbook


class PasswordTools:
    def __init__(self):
        self.__characters = string.ascii_letters + string.digits + string.punctuation
        self.__key = None
    
    @staticmethod
    def __generate_password(length: int, characters: str) -> str:
        try:
            if characters == '1' or characters == "":
                characters = string.ascii_letters + string.digits + string.punctuation

            elif characters == '2':
                characters = string.ascii_letters

            elif characters == '3':
                characters = string.digits

            elif characters == '4':
                characters = string.punctuation

            elif characters == '5':
                characters = string.ascii_letters + string.digits

            elif characters == '6':
                characters = string.ascii_letters + string.punctuation

            elif characters == '7':
                characters = string.digits + string.punctuation

            elif characters == '8':
                characters = string.printable + ''.join(chr(i) for i in range(0x80, 0x110000))

            password = ''.join(secrets.choice(characters) for _ in range(length))

            return password

        except Exception as e:
            print(f"\n\033[3m!✶ Error: \033[38;5;200m{e}\033[0m")

    @staticmethod
    def generate_password(length: int, characters: str) -> str:
        return PasswordTools.__generate_password(length, characters)

    @staticmethod
    def check_strength(password: str) -> str:
        try:
            result = zxcvbn.zxcvbn(password)
            score = result["score"]

            if score == 0:
                return "Very Weak"
            
            elif score == 1:
                return "Weak"
            
            elif score == 2:
                return "Medium"
            
            elif score == 3:
                return "Strong"
            
            elif score == 4:
                return "Very Strong"
            
        except Exception as e:
            print(f"\n\033[3m!✶ Error: \033[38;5;200m{e}\033[0m")

    @staticmethod
    def __save_to_xslx(passlen: int, password: str, hashed_password: str, encrypted_password: bytes, crypto_key: bytes):
        try:
            try:
                wb = load_workbook('passwords.xlsx')
                ws = wb.active

            except FileNotFoundError:
                wb = Workbook()
                ws = wb.active
                ws['A1'] = 'Date & Time'
                ws['B1'] = 'Length'
                ws['C1'] = 'Passwords'
                ws['D1'] = 'Hashed Values'
                ws['E1'] = 'Encrypted Passwords'
                ws['F1'] = 'Crypto Keys'

            row = ws.max_row + 1
            ws[f'A{row}'] = dt.now().strftime('%Y-%m-%d %H:%M:%S')

            if passlen is not None:
                ws[f'B{row}'] = passlen

            if password is not None:
                ws[f'C{row}'] = password

            if hashed_password is not None:
                ws[f'D{row}'] = hashed_password

            if encrypted_password is not None:
                ws[f'E{row}'] = encrypted_password

            if crypto_key is not None:
                ws[f'F{row}'] = crypto_key

            wb.save('passwords.xlsx')

            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length

        except Exception as e:
            print(f"\n\033[3m!✶ Error: \033[38;5;200m{e}\033[0m")

    @staticmethod
    def save_to_xslx(passlen: int, password: str, hashed_password: str, encrypted_password: bytes, crypto_key: bytes):
        return PasswordTools.__save_to_xslx(passlen, password, hashed_password, encrypted_password, crypto_key)

    @staticmethod
    def __save_to_txt(password: str):
        try:
            with open('passwords.txt', 'a') as txtfile:
                txtfile.write(f'{password}\n')

        except Exception as e:
            print(f"\n\033[3m!✶ Error: \033[38;5;200m{e}\033[0m")

    @staticmethod
    def save_to_txt(password: str):
        return PasswordTools.__save_to_txt(password)

    @staticmethod
    def __hash_password(password: str) -> str:
        try:
            return passlib.hash.pbkdf2_sha256.using(salt_size=32, rounds=2**20).hash(password)
        
        except Exception as e:
            print(f"\n\033[3m!✶ Error: \033[38;5;200m{e}\033[0m")
    
    @staticmethod
    def hash_password(password: str) -> str:
        return PasswordTools.__hash_password(password)

    @staticmethod
    def __verify_hash(password: str, hashed_password: str) -> bool:
        try:
            return passlib.hash.pbkdf2_sha256.verify(password, hashed_password)
        
        except ValueError:
            return False
        
        except Exception as e:
            print(f"\n\033[3m!✶ Error: \033[38;5;200m{e}\033[0m")

    @staticmethod
    def verify_hash(password: str, hashed_password: str) -> bool:
        return PasswordTools.__verify_hash(password, hashed_password)

    def __encrypt_password(self, password: str) -> bytes:
        try:
            self.__key = Fernet.generate_key()  #base64
            cipher_suite = Fernet(self.__key)
            encrypted_password = cipher_suite.encrypt(password.encode())

            with open('key.key', 'ab') as key_file:
                key_file.write(self.__key + b"\n")

            return encrypted_password
        
        except Exception as e:
            print(f"\n\033[3m!✶ Error: \033[38;5;200m{e}\033[0m")

    def encrypt_password(self, password: str) -> bytes:
        return self.__encrypt_password(password)

    @staticmethod
    def __decrypt_password(encrypted_password: bytes) -> str:
        with open('key.key', 'rb') as key_file:
            keys = key_file.read().split(b'\n')

        for key in keys:
            cipher_suite = Fernet(key)
            try:
                decrypted_password = cipher_suite.decrypt(encrypted_password).decode()
                return decrypted_password
            
            except Exception:
                pass
            
        return "Password decryption failed"

    @staticmethod
    def decrypt_password(encrypted_password: bytes) -> str:
        return PasswordTools.__decrypt_password(encrypted_password)

    @staticmethod
    def __fetch_crypto_keys(encrypted_password: bytes) -> str:
        with open('key.key', 'rb') as kf:
            keys = kf.read().split(b'\n')
        
        for key in keys:
            cipher_suite = Fernet(key)
            try:
                decrypted_password = cipher_suite.decrypt(encrypted_password).decode()
                return key.decode()
            
            except Exception:
                pass
            
        return "Crypto key not found"

    @staticmethod
    def fetch_crypto_keys(encrypted_password: bytes) -> str:
        return PasswordTools.__fetch_crypto_keys(encrypted_password)

    def __run(self):
        print("✵✵✵ Password Tools Manager ✵✵✵")
        while True:
            try:
                print("\n1. Generate Password\n2. Hash Password\n3. Verify Hash\n4. Encrypt Password\n5. Decrypt Password\n6. Retrieve Crypto Key\n0. Exit")
                action = input("\nResponse: ")

                if action == "0":
                    break

                if action == "1":
                    passlen = input("\bPasslen: ")

                    if passlen == "":
                        break
                    
                    if not passlen.isdigit():
                        raise ValueError("\nInvalid password length. Please enter a positive integer.")
                    
                    passlen = int(passlen)

                    if passlen == 0:
                        break
                    
                    elif passlen < 0:
                        raise ValueError("\nInvalid password length. Please enter a positive integer.")

                    print("\n1. All Mixed\n2. Letters\n3. Digits\n4. Special Characters\n5. Letters + Digits")
                    print("6. Letters + Special Characters\n7. Digits + Special Characters\n8. ASCII + Unicode Combinations")

                    characters = input("\nChoice: ")

                    password = PasswordTools.generate_password(passlen, characters)

                    if password is None:
                        print("\nPlease enter a valid password length.")
                        continue
                    
                    strength = PasswordTools.check_strength(password)
                    PasswordTools.save_to_xslx(passlen, password, None, None, None)
                    PasswordTools.save_to_txt(password)

                    print(f"\n\033[38;5;203;3mPassword: \033[38;5;27m{password}\033[0m")
                    print(f"\033[38;5;158;3mStrength: \033[38;5;229m{strength}\033[0m")

                    while True:
                        try:
                            response = input("\nTry another one? (y/n): ")
                            if response.upper() not in ['Y', 'N', ""]:
                                raise ValueError("\nInvalid response. Please enter Y or N.")
                            
                        except ValueError as ve:
                            print(f"\n\033[3m!✶ Error: \033[38;5;200m{ve}\033[0m")
                            continue
                        
                        if response.upper() == 'N' or response == "":
                            return
                        else:
                            break

                if action == "2":
                    try:
                        passwort = input("Submit Password: ")
                        hashed_password = PasswordTools.hash_password(passwort)
                        PasswordTools.save_to_xslx(len(passwort), passwort, hashed_password, None, None)
                        print(f"\nHashed Value: \033[38;5;4m{hashed_password}\033[0m")

                    except Exception as e:
                        print(f"\n\033[3m!✶ Error: \033[38;5;200m{e}\033[0m")

                    while True:
                        try:
                            response = input("\nTry another one? (y/n): ")
                            if response.upper() not in ['Y', 'N', ""]:
                                raise ValueError("\nInvalid response. Please enter Y or N.")
                            
                        except ValueError as ve:
                            print(f"\n\033[3m!✶ Error: \033[38;5;200m{ve}\033[0m")
                            continue
                        
                        if response.upper() == 'N' or response == "":
                            return
                        else:
                            break

                if action == "3":
                    try:
                        passwort = input("Password: ")
                        hash_verification = input("Verify Hash: ")
                        verified_hash = PasswordTools.__verify_hash(passwort, hash_verification)

                        if verified_hash:
                            print("\n\033[38;5;82mVerified!\033[0m")
                        else:
                            print("\n\033[38;5;202mVerification failed!\033[0m")

                    except Exception as e:
                        print(f"\n\033[3m!✶ Error: \033[38;5;200m{e}\033[0m")

                    while True:
                        try:
                            response = input("\nTry another one? (y/n): ")
                            if response.upper() not in ['Y', 'N', ""]:
                                raise ValueError("\nInvalid response. Please enter Y or N.")
                            
                        except ValueError as ve:
                            print(f"\n\033[3m!✶ Error: \033[38;5;200m{ve}\033[0m")
                            continue
                        
                        if response.upper() == 'N' or response == "":
                            return
                        else:
                            break

                if action == "4":
                    try:
                        enc_password = input("Password: ")
                        encrypted_password = self.encrypt_password(enc_password)
                        PasswordTools.save_to_xslx(len(enc_password), enc_password, None, encrypted_password, self.__key)
                        print(f"\nEncrypted Password: \033[38;5;6m{encrypted_password}\033[0m")

                    except Exception as e:
                        print(f"\n\033[3m!✶ Error: \033[38;5;200m{e}\033[0m")

                    while True:
                        try:
                            response = input("\nTry another one? (y/n): ")
                            if response.upper() not in ['Y', 'N', ""]:
                                raise ValueError("\nInvalid response. Please enter Y or N.")
                            
                        except ValueError as ve:
                            print(f"\n\033[3m!✶ Error: \033[38;5;200m{ve}\033[0m")
                            continue
                        
                        if response.upper() == 'N' or response == "":
                            return
                        else:
                            break

                if action == "5":
                    try:
                        dec_password = input("Encrypted Password: ")

                        # Convert the user input to a byte string
                        if dec_password.startswith("b'") and dec_password.endswith("'"):
                            dec_password = dec_password[2:-1].encode()
                        elif dec_password.startswith('b"') and dec_password.endswith('"'):
                            dec_password = dec_password[2:-1].encode()

                        decrypted_password = PasswordTools.decrypt_password(dec_password)
                        print(f"\nDecrypted Password: \033[38;5;9m{decrypted_password}\033[0m")

                    except Exception as e:
                        print(f"\n\033[3m!✶ Error: \033[38;5;200m{e}\033[0m")

                    while True:
                        try:
                            response = input("\nTry another one? (y/n): ")
                            if response.upper() not in ['Y', 'N', ""]:
                                raise ValueError("\nInvalid response. Please enter Y or N.")
                            
                        except ValueError as ve:
                            print(f"\n\033[3m!✶ Error: \033[38;5;200m{ve}\033[0m")
                            continue
                        
                        if response.upper() == 'N' or response == "":
                            return
                        else:
                            break

                if action == "6":
                    try:
                        retrieve_crypto_key = input("Encrypted Password: ")

                        # Convert the user input to a byte string
                        if retrieve_crypto_key.startswith("b'") and retrieve_crypto_key.endswith("'"):
                            retrieve_crypto_key = retrieve_crypto_key[2:-1].encode()
                        elif retrieve_crypto_key.startswith('b"') and retrieve_crypto_key.endswith('"'):
                            retrieve_crypto_key = retrieve_crypto_key[2:-1].encode()

                        crypto_key = PasswordTools.fetch_crypto_keys(retrieve_crypto_key)
                        print(f"\nCrypto Key: \033[38;5;250m{crypto_key}\033[0m")

                    except Exception as e:
                        print(f"\n\033[3m!✶ Error: \033[38;5;200m{e}\033[0m")

                    while True:
                        try:
                            response = input("\nTry another one? (y/n): ")
                            if response.upper() not in ['Y', 'N', ""]:
                                raise ValueError("\nInvalid response. Please enter Y or N.")
                            
                        except ValueError as ve:
                            print(f"\n\033[3m!✶ Error: \033[38;5;200m{ve}\033[0m")
                            continue
                        
                        if response.upper() == 'N' or response == "":
                            return
                        else:
                            break

            except ValueError as ve:
                return f"\n\033[3m!✶ Error: \033[38;5;200m{ve}\033[0m"

            except Exception as e:
                print(f"\n\033[3m!✶ Error: \033[38;5;200m{e}\033[0m")

    def run(self):
        self.__run()


if __name__ == "__main__":
    passtools = PasswordTools()
    passtools.run()