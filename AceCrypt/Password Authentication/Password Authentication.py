
import logging
import openpyxl
import tkinter as tk
from tkinter import messagebox
from cryptography.fernet import Fernet


logging.basicConfig(filename='authentication.log', level=logging.INFO, format='%(asctime)s:%(levelname)s:%(message)s')


class AuthenticationApp(tk.Tk):
    def __init__(self, file_name, key_file_name):
        super().__init__()
        self.title("Authentication System")
        
        self.__file_name = file_name
        self.__key_file_name = key_file_name
        
        try:
            with open(self.__key_file_name, 'rb') as f:
                self.__key = f.read()

        except FileNotFoundError:
            self.__key = Fernet.generate_key()
            with open(self.__key_file_name, 'wb') as f:
                f.write(self.__key)

        self.__database = self.__load_database()

        self.__username_label = tk.Label(self, text="Username:")
        self.__username_entry = tk.Entry(self)
        
        self.__password_label = tk.Label(self, text="Password:")
        self.__password_entry = tk.Entry(self, show="*")
        
        self.__register_button = tk.Button(self, text="Register", command=self.register)
        self.__authenticate_button = tk.Button(self, text="Authenticate", command=self.authenticate)
        
        self.__username_label.grid(row=0, column=0)
        self.__username_entry.grid(row=0, column=1)
        
        self.__password_label.grid(row=1, column=0)
        self.__password_entry.grid(row=1, column=1)
        
        self.__register_button.grid(row=2, column=0)
        self.__authenticate_button.grid(row=2, column=1)

    def __load_database(self):
        try:
            wb = openpyxl.load_workbook(self.__file_name)
            sheet = wb.active
            database = {}
            for row in sheet.iter_rows(values_only=True):
                if row[0] is not None and row[1] is not None:
                    database[row[0]] = row[1]
            return database

        except Exception as e:
            logging.error(f"An error occurred while loading the database: {e}")
            return {}

    def __save_database(self):
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            for i, (username, password) in enumerate(self.__database.items()):
                sheet.cell(row=i+1, column=1).value = username
                sheet.cell(row=i+1, column=2).value = password
            wb.save(self.__file_name)

        except Exception as e:
            logging.error(f"An error occurred while saving the database: {e}")

    def __encrypt_password(self,password):
        f = Fernet(self.__key)
        encrypted_password = f.encrypt(password.encode())
        return encrypted_password.decode()

    def __decrypt_password(self,encrypted_password):
        f = Fernet(self.__key)
        decrypted_password = f.decrypt(encrypted_password.encode())
        return decrypted_password.decode()

    def register(self):
        username = self.__username_entry.get()
        password = self.__password_entry.get()
        
        if username in self.__database:
            messagebox.showerror("Error", "This username is already taken. Please choose a different username.")
            return
        
        encrypted_password = self.__encrypt_password(password)
        self.__database[username] = encrypted_password
        self.__save_database()
        
        logging.info(f"New user registered: {username}")
        
        messagebox.showinfo("Success", "User registered successfully.")

    def authenticate(self):
        username = self.__username_entry.get()
        password = self.__password_entry.get()
        
        if username in self.__database:
            encrypted_password = self.__database[username]
            decrypted_password = self.__decrypt_password(encrypted_password)
            if password == decrypted_password:
                logging.info(f"User authenticated: {username}")
                messagebox.showinfo("Success", "Verified")
                return
        
        logging.warning(f"Failed authentication attempt: {username}")
        
        messagebox.showerror("Error", "Invalid credentials")

    def __run(self):
      print("Running the Authentication App")
      print(f"File name: {self.file_name}")
      print(f"Key file name: {self.key_file_name}")
      print(f"Database: {self.database}")


if __name__ == "__main__":
    file_name = "database.xlsx"
    key_file_name = "key.key"
    
    app = AuthenticationApp(file_name, key_file_name)
    app.mainloop()
