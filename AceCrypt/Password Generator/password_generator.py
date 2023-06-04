
import string
import secrets
from openpyxl import Workbook, load_workbook
from datetime import datetime


class PasswordGenerator:
    def __init__(self):
        try:
            self.__characters = string.ascii_letters + string.digits + string.punctuation
        except Exception as e:
            print(f"\n\u001b[3m** An error occurred: \u001b[38;5;200m{e}\u001b[0m")

    @staticmethod
    def __generate_password(length: int, characters: str) -> str:
        try:
            if characters == "" or characters == '1':
                characters = string.ascii_letters + string.digits + string.punctuation
            elif characters == 'letters' or characters == '2':
                characters = string.ascii_letters
            elif characters == 'digits' or characters == '3':
                characters = string.digits
            elif characters == 'punctuation' or characters == '4':
                characters = string.punctuation
            elif characters == 'letters_digits' or characters == '5':
                characters = string.ascii_letters + string.digits
            elif characters == 'letters_punctuation' or characters == '6':
                characters = string.ascii_letters + string.punctuation
            elif characters == 'digits_punctuation' or characters == '7':
                characters = string.digits + string.punctuation
            elif characters == 'ascii unicode' or characters == "8":
                characters = string.printable + ''.join(chr(i) for i in range(0x80, 0x110000))

            password = ''.join(secrets.choice(characters) for _ in range(length))
            return password

        except Exception as e:
            print(f"\n\u001b[3m** An error occurred: \u001b[38;5;200m{e}\u001b[0m")

    @staticmethod
    def generate_password(length: int, characters: str) -> str:
        return PasswordGenerator.__generate_password(length, characters)

    @staticmethod
    def __save_to_excel(password: str, passlen: int):
        try:
            try:
                wb = load_workbook('passwords.xlsx')
                ws = wb.active
            except FileNotFoundError:
                wb = Workbook()
                ws = wb.active
                ws['A1'] = 'Date & Time'
                ws['B1'] = 'Length'
                ws['C1'] = 'Password'
            row = ws.max_row + 1
            ws[f'A{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ws[f'B{row}'] = passlen
            ws[f'C{row}'] = password
            wb.save('passwords.xlsx')
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length
        except Exception as e:
            print(f"\n\u001b[3m** An error occurred: \u001b[38;5;200m{e}\u001b[0m")

    @staticmethod
    def save_to_excel(password: str, passlen: int):
        PasswordGenerator.__save_to_excel(password, passlen)

    @staticmethod
    def __save_to_text_file(password: str):
        try:
            with open('passwords.txt', 'a') as textfile:
                textfile.write(f'{password}\n')
        except Exception as e:
            print(f"\n\u001b[3m** An error occurred: \u001b[38;5;200m{e}\u001b[0m")

    @staticmethod
    def save_to_text_file(password: str):
        PasswordGenerator.__save_to_text_file(password)

    @staticmethod
    def __run():
        print("✵✵✵ Welcome to Password Generator ✵✵✵")
        while True:
            try:
                passlen_str = input("\nPasslen: ")
                if passlen_str == "q" or passlen_str == '':
                    break
                if not passlen_str.isdigit():
                    raise ValueError("\nInvalid password length. Please enter a positive integer.")
                passlen = int(passlen_str)

                if passlen == 0:
                    break
                elif passlen < 0:
                    raise ValueError("\nInvalid password length. Please enter a positive integer.")
            except ValueError as e:
                print(f"\n\u001b[3m** An error occurred: \u001b[38;5;200m{e}\u001b[0m")
                continue

            print("\nOptions:\n1. All mixed\n2. Letters\n3. Digits\n4. Special Characters\n5. Letters + Digits")
            print("6. Letters + Special Characters\n7. Digits + Special Characters\n8. ASCII + Unicode Combinations")

            characters = input("\nChoice: ")
            password = PasswordGenerator.generate_password(passlen, characters)
            print(f"\n\u001b[38;5;3;3mPassword: \u001b[38;5;27m{password}\u001b[0m")
            PasswordGenerator.save_to_excel(password, passlen)
            PasswordGenerator.save_to_text_file(password)

            while True:
                try:
                    response = input("\nTry another one? (y/n): ")
                    if response.upper() not in ['Y', 'N', ""]:
                        raise ValueError("\nInvalid response. Please enter Y or N.")
                except ValueError as e:
                    print(f"\n\u001b[3m** An error occurred: \u001b[38;5;200m{e}\u001b[0m")
                    continue
                if response.upper() == 'N' or response == "":
                    exit(0)
                else:
                    break
        print("\nThanks for using the Password Generator!")

    @staticmethod
    def run():
        PasswordGenerator.__run()


if __name__ == "__main__":
    passgen = PasswordGenerator()
    passgen.run()
