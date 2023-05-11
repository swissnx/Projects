
import string
import secrets
from openpyxl import Workbook, load_workbook


# def generate_password(length: int) -> str:
#     characters = string.ascii_letters + string.digits + string.punctuation
#     password = ''.join(secrets.choice(characters) for _ in range(length))
#     return password

def generate_password(length: int, characters: str) -> str:
    if characters == "" or characters == '1':
        characters = string.ascii_letters + string.digits + string.punctuation
    elif characters == 'letters' or characters == '2':
        characters = string.ascii_letters
    elif characters == 'digits' or characters == '3':
        characters = string.digits
    elif characters == 'punctuation' or characters == '4':
        characters = string.punctuation
    elif characters == 'letters_digits' or characters == '5':
        characters = string.ascii_letters + string.digits
    elif characters == 'letters_punctuation' or characters == '6':
        characters = string.ascii_letters + string.punctuation
    elif characters == 'digits_punctuation' or characters == '7':
        characters = string.digits + string.punctuation

    password = ''.join(secrets.choice(characters) for _ in range(length))
    return password


def save_to_excel(passwort: str):
    try:
        wb = load_workbook('passwords.xlsx')
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Password'
    row = ws.max_row + 1
    ws[f'A{row}'] = passwort
    wb.save('passwords.xlsx')


def save_to_text_file(passwortxt: str):
    with open('passwords.txt', 'a') as textfile:
        textfile.write(f'{passwortxt}\n')



if __name__ == '__main__':
    print("✵✵✵ Welcome to Password Generator ✵✵✵")
    while True:
        try:
            passlen_str = input("\nPasslen: ")
            if passlen_str == "q" or passlen_str == '':
                break
            if not passlen_str.isdigit():
                raise ValueError("\nInvalid password length. Please enter a positive integer.")
            passlen = int(passlen_str)

            if passlen == 0:
                break
            elif passlen < 0:
                raise ValueError("\nInvalid password length. Please enter a positive integer.")
        except ValueError as e:
            print(f"\n\u001b[3m** Problem is: \u001b[38;5;196m{e}\u001b[0m")
            continue

        print("\nOptions:")
        print("1. All mixed")
        print("2. Letters only")
        print("3. Digits only")
        print("4. Special Characters only")
        print("5. Letters + Digits")
        print("6. Letters + Special Characters")
        print("7. Digits + Special Characters")
        
        characters = input("\nChoice: ")
        password = generate_password(passlen, characters)
        print(f"\n\u001b[38;5;3;3mPassword: \u001b[38;5;27m{password}\u001b[0m")
        save_to_excel(password)
        save_to_text_file(password)

        while True:
            try:
                response = input("\nTry another one? (y/n): ")
                if response.upper() not in ['Y', 'N', ""]:
                    raise ValueError("\nInvalid response. Please enter Y or N.")
            except ValueError as e:
                print(f"\n\u001b[3m** Problem is: \u001b[38;5;196m{e}\u001b[0m")
                continue
            if response.upper() == 'N' or response == "":
                exit(0)
            else:
                break

    print("\nThanks for using the Password Generator. Goodbye!")
