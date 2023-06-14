
import datetime
import openpyxl
from openpyxl import Workbook


class RefNumGen:
    def __init__(self):
        self.__ref_num = None
        self.__custom_digits = 0
        self.__file_name = "ref_nums.xlsx"
        self.__load_ref_nums()

    def __load_ref_nums(self):
        try:
            wb = openpyxl.load_workbook(self.__file_name)
            sheet = wb.active
            max_row = sheet.max_row
            last_ref_num = sheet.cell(row=max_row, column=2).value
            self.__custom_digits = int(last_ref_num[4:]) + 1
        except FileNotFoundError:
            wb = Workbook()
            sheet = wb.active
            sheet.append(["Date", "Reference Numbers", "Notes"])
            wb.save(self.__file_name)

    def __generate_ref_num(self):
        try:
            yy_mm = f"{datetime.datetime.now():%y%m}"
            custom_digits = f"{self.__custom_digits:05d}"
            self.__ref_num = f"{yy_mm}{custom_digits}"
            self.__custom_digits += 1
        except Exception as e:
            print(f"\n\u001b[3m** An error occurred: \u001b[38;5;200m{e}\u001b[0m")

    def get_ref_num(self):
        try:
            self.__generate_ref_num()
            return self.__ref_num
        except Exception as e:
            print(f"\n\u001b[3m** An error occurred: \u001b[38;5;200m{e}\u001b[0m")

    def __save_ref_num(self, notes=""):
        try:
            date_str = f"{datetime.datetime.now():%Y-%m-%d}"
            wb = openpyxl.load_workbook(self.__file_name)
            sheet = wb.active
            sheet.append([date_str, self.__ref_num, notes])
            wb.save(self.__file_name)
        except Exception as e:
            print(f"\n\u001b[3m** An error occurred: \u001b[38;5;200m{e}\u001b[0m")

    def __verify_ref_num(self, ref_num):
        try:
            wb = openpyxl.load_workbook(self.__file_name)
            sheet = wb.active
            for row in sheet.iter_rows(values_only=True):
                if row[1] == ref_num:
                    return True
            return False
        except Exception as e:
            print(f"\n\u001b[3m** An error occurred: \u001b[38;5;200m{e}\u001b[0m")

    def verify_ref_num(self, ref_num):
        return self.__verify_ref_num(ref_num)

    def __run(self):
        while True:
            try:
                print("\nOptions:\n1. Generate\n2. Verify\n3. Exit")
                option = input("\nChoice: ")
                if option == "1":
                    print(f"Reference Number: {self.get_ref_num()}")
                    notes = input("Notes: ")
                    self.__save_ref_num(notes)
                elif option == "2":
                    ref_num = input("Enter reference number: ")
                    if self.__verify_ref_num(ref_num):
                        print("Valid!")
                    else:
                        print("Invalid!")
                elif option == "3" or option == "":
                    break
                else:
                    print("Invalid option.")
            except Exception as e:
                print(f"\n\u001b[3m** An error occurred: \u001b[38;5;200m{e}\u001b[0m")

    def run(self):
        self.__run()


if __name__ == "__main__":
    rng = RefNumGen()
    rng.run()
